#LIBRERIAS
from collections import Counter
import math
from tkinter import *
from tkinter import messagebox
from PIL import ImageTk, Image
from tkinter.messagebox import *
from tkinter.ttk import Combobox
import webbrowser
from tkinter.filedialog import askopenfilename
import re
from matplotlib import pyplot as plt
import numpy as np
import pyperclip
from scipy import stats as st 
import statistics as stat
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import ttk

class app(Frame):
    def __init__(self, master=None, width = None, height= None):
        super().__init__(master, width = width ,height= height)
        self.tema(1)
        self.master = master
        self.logotipoI = ImageTk.PhotoImage(Image.open("img/logotipoB.png").resize((round(width/2.25),round(height/3))))
        self.iconoI = ImageTk.PhotoImage(Image.open("img/icono.ico").resize((60,60)))
        self.place(relx=0,rely=0.9, relheight =1, relwidth= 1)
        self.pantalla_principal()


    def pantalla_principal(self):
        self.borrar()
        self.frameI = Frame(background="white")
        self.imagen_logotipo = Label(self.frameI,image=self.logotipoI).place(relx=0.52,rely=0.4, relheight = 0.3, relwidth= 0.4,anchor=CENTER)
        self.icono = Label(self.frameI, image=self.iconoI, background="white").place(relx=0.34,rely=0.38, relheight = 0.1, relwidth= 0.05,anchor=CENTER)
        self.boton = Button(self.frameI, text="COMENZAR", font=("Segoe UI",round(widthroot*1/100),"bold"), background="gold2", fg="gray9",  cursor="hand2", border=13, borderwidth=10,command= self.crearcombos).place(relx=0.5,rely=0.7, anchor=CENTER, relheight=0.055,relwidth=0.10)
        
        self.frameI.place(relx=0, rely=0.05,relheight=1, relwidth=1)
        

    def cargarA(self):
        self.crearcombos()
        try:
            self.framecon.destroy()
        except:pass
        self.sin_intervalo()
        self.check_con_inter.deselect()
        self.check_sin_inter.select()
        try: 
                if self.controlcheckB.get() == "conI":
                    self.colorletracheck(2)
                else: self.colorletracheck(1)
        except:pass
        self.entrada_datos.delete(1.0,END)
        self.file = askopenfilename()
        with open(self.file,"r") as texto:
            self.lectura = texto.read()
            self.entrada_datos.insert(1.0,self.lectura)

    def cargarB(self):
        self.crearcombos()
        try:
            self.framesin.destroy()
        except:pass
        self.con_intervalo()
        self.check_sin_inter.deselect()
        self.check_con_inter.select()
        try: 
                if self.controlcheckB.get() == "conI":
                    self.colorletracheck(2)
                else: self.colorletracheck(1)
        except:pass
        self.entrada_sacarI.delete(1.0,END)
        self.file = askopenfilename()
        with open(self.file,"r") as texto:
            self.lectura = texto.read()
            self.entrada_sacarI.insert(1.0,self.lectura)

    def borrar(self):   
        self.list=self.place_slaves()
        for k in self.list:
            k.destroy()
    
    def salir(self):
        salir = askquestion(title="STADISKIS", message="¿Desea salir de Stadiskis?")
        if salir == "yes":
            self.quit()


    def volver_datos(self):
        self.pantalla_principal()

    def boton_volver(self):
         self.boton_volver_datos= Button(self,text="VOLVER",font=("Segoe UI",round(widthroot*0.8/100),"bold"), cursor="hand2", width=5,command=self.volver_datos, borderwidth=5).place(relx=0.025,rely=0.95, anchor=CENTER, relheight=0.04,relwidth=0.05)

    def colorletraboton(self, num):
                if num == 2:
                    self.radio2.config(foreground="gold4")
                    self.radio.config(foreground=self.letra)
                else:
                    self.radio2.config(foreground=self.letra)
                    self.radio.config(foreground="gold4")
    def creador_Rboton(self, frame, x, y ):
            
            self.valor_r = IntVar()
            self.valor_r.set(value=1)
            self.radio = Radiobutton(frame,text="Población", font=("",round(widthroot*1.1/100)), variable=self.valor_r,value=1, background=self.colorlabel, foreground=self.letra)
            self.radio2 = Radiobutton(frame,text="Muestra",font=("",round(widthroot*1.1/100)),variable=self.valor_r,value=2, background=self.colorlabel, foreground=self.letra)
            if x == 0:
                self.radio.place(relx=0.2,rely=0.38,  relheight=0.03,relwidth=0.2)
                self.radio2.place(relx=0.4,rely=0.38, relheight=0.03,relwidth=0.2)
            else:
                try:
                    self.radio2.place(relx=x+0.20,rely=y, relheight=0.04,relwidth=0.2)
                    self.radio.place(relx=x+0.01,rely=y,  relheight=0.04,relwidth=0.23)
                    
                except:pass
                
            self.radio2.bind("<Button-1>", lambda i:self.colorletraboton(2) )
            self.radio.bind("<Button-1>", lambda i: self.colorletraboton(1))
            self.colorletraboton(1)
            
    
    def tema(self, color):
        self.pantalla = ""
        self.letra = ""
        self.colorlabel= ""
        self.fondotext = ""
        self.letratext = ""
        self.fondotext=""
        self.letratext = ""
        self.fondoCD =""
        self.fondocheck = ""
        self.letracheck = ""
        def validar_modo():
            validacion = color
            if validacion == 1:
                modo_oscuro()
            else:modo_claro()
        def modo_claro():
            self.pantalla = "gray95"
            self.letra = "gray9"
            self.colorlabel ="gray95"
            self.fondotext = "white"
            self.letratext = "gray9"
            self.config(background="gray95")
            self.fondoCD = "gold2"
            self.fondocheck = "gray95"
            self.letracheck = "gold4"
            try:
                self.label1C.config(background="gray95", foreground="gray9") 
                self.resultado.config(background="gray95", foreground="gray9")
                self.salida.config(background="white", foreground="gray9")
                self.radio.config(background="gray95", foreground="gray9")
                self.radio2.config(background="gray95", foreground="gray9")
                self.check_sin_inter.config(background="gray95", foreground="gray9")
                self.check_con_inter.config(background="gray95", foreground="gray9")
            except: pass
            try:
                self.datos.config(background="gray95", foreground="gray9")
                self.operacion.config(background="gray95", foreground="gray9")
                self.entrada_datos.config(background="white", foreground="gray9")
                self.framesin.config(background="gray95")
                self.labelsin.config(background="gray95", foreground="gray9")
            except:pass
            try:
                self.framecon.config(background="gray95")
                self.label_sacarI.config(background="gray95", foreground="gray9")
                self.entrada_sacarI.config(background="white", foreground="gray9")
                self.frecuenciaL.config(background="gray95",foreground="gray9")
                self.intervalo_label.config(background="gray95",foreground="gray9")
                self.labelcon.config(background="gray95", foreground="gray9")
                self.operacion.config(background="gray95", foreground="gray9")
                self.frecuenciaL2.config(background="gray95", foreground="gray9")
                self.frecuenciaL1.config(background="gray95", foreground="gray9")
                self.intervalo_label2.config(background="gray95", foreground="gray9")
                self.intervalo_label1.config(background="gray95", foreground="gray9")
            except:pass
            try: 
                if self.controlcheckB.get() == "conI":
                    self.colorletracheck(2)
                else: self.colorletracheck(1)
            except:pass
            try:
                if self.valor_r == 1:
                    self.colorletraboton(2)
                else:self.colorletraboton(1)
            except:pass

        def modo_oscuro():
            self.fondocheck = "gray9"
            self.letracheck = "gold4"
            self.pantalla = "gray9"
            self.letra = "gray70"
            self.colorlabel = "gray12"
            self.fondotext = "gray18"
            self.letratext = "gray70"
            self.config(background="gray12")
            self.fondoCD = "gray9"
            try:
                self.label1C.config(background="gray9", foreground="gray70")
                self.resultado.config(background="gray12", foreground="gray70")
                self.salida.config(background="gray18", foreground="gray70")  
                self.radio.config(background="gray12", foreground="gray70")
                self.radio2.config(background="gray12", foreground="gray70")
                self.check_con_inter.config(background="gray9", foreground="gray70")
                self.check_sin_inter.config(background="gray9", foreground="gray70")
            except:pass
            try:
                self.framesin.config(background="gray12")
                self.entrada_datos.config(background="gray18", foreground="gray70")
                self.labelsin.config(background="gray12", foreground="gray70")
                self.datos.config(background="gray12", foreground="gray70")
                self.operacion.config(background="gray12", foreground="gray70")
            except:pass
            try:
                self.framecon.config(background="gray12")
                self.label_sacarI.config(background="gray12", foreground="gray70")
                self.entrada_sacarI.config(background="gray18", foreground="gray70")
                self.frecuenciaL.config(background="gray12",foreground="gray70")
                self.intervalo_label.config(background="gray12",foreground="gray70")
                self.labelcon.config(background="gray12", foreground="gray70")
                self.operacion.config(background="gray12", foreground="gray70")
                self.frecuenciaL2.config(background="gray12", foreground="gray70")
                self.frecuenciaL1.config(background="gray12", foreground="gray70")
                self.intervalo_label2.config(background="gray12", foreground="gray70")
                self.intervalo_label1.config(background="gray12", foreground="gray70")
            except:pass
            try: 
                if self.controlcheckB.get() == "conI":
                    self.colorletracheck(2)
                else: self.colorletracheck(1)
            except:pass
            try:
                if self.valor_r == 1:
                    self.colorletraboton(2)
                else:self.colorletraboton(1)
            except:pass
        validar_modo()

    def colorletracheck(self, num):
                if num == 2:
                    self.check_con_inter.config(foreground="gold4")
                    self.check_sin_inter.config(foreground=self.letra)
                else:
                    self.check_con_inter.config(foreground=self.letra)
                    self.check_sin_inter.config(foreground="gold4")
    def crearcombos(self):
        self.borrar()
        self.frameI.destroy()
        self.boton_volver()
        self.controlcheckB= StringVar()
        
        self.separ2 = Frame(self, background="gray20").place(relx=0, rely=0.11 , relheight=0.001, relwidth=1)
        self.separ = Frame(self, background="gray20").place(relx=0.55, rely=0.11 , relheight=0.95, relwidth=0.001)

        self.label1C = Label(self,text="SELECCIONE QUE TIPO DE DATOS DESEA CARGAR",
                             font=("Segoe UI",round(widthroot*1.5/100),"bold"),
                             background=self.pantalla,
                             foreground=self.letra)
        self.check_con_inter = Checkbutton(root,
					    text="Con intervalo",
					    variable=self.controlcheckB,
					    onvalue="conI",
					    offvalue="n",
                        font=("Segoe UI",round(widthroot*1.5/100), "bold"),
                        foreground=self.letra,
                        command=self.con_intervalo,
                        background=self.fondocheck)
        self.check_sin_inter = Checkbutton(root,
					    text="Sin intervalo",
					    variable=self.controlcheckB,
					    onvalue="sinI",
					    offvalue="n",
                        font=("Segoe UI",round(widthroot*1.5/100), "bold"),
                        foreground=self.letra,
                        command=self.sin_intervalo,
                        background=self.fondocheck)
        self.check_con_inter.deselect()
        self.colorletracheck(1)

        self.controlcheckB.set("sinI")
        
        
        self.scrol = Scrollbar(self, cursor="hand2",
                               orient= VERTICAL,
                               activebackground="gold2")
        self.salida= Text(self,background=self.fondotext,
                          foreground=self.letratext,
                          font=("Segoe UI",round(widthroot*1.1/100)),
                          yscrollcommand=self.scrol.set)
        self.scrol.config(command=self.salida.yview)
        self.borrar_salida = Button(self,text="Borrar",
                                    cursor="hand2",
                                    command= lambda :self.salida.delete(1.0,END)
                                    )
        self.resultado = Label(self, text="Resultado:",
                               fg=self.letra,
                               font=("Segoe UI",round(widthroot*1.1/100)),
                               background=self.colorlabel)
        

        self.label1C.place(relx=0.5,rely=0.035, anchor=CENTER, relheight=0.15,relwidth=1)
        self.check_con_inter.place(relx=0.6,rely=0.11, anchor=CENTER, relheight=0.03,relwidth=0.2)
        self.check_sin_inter.place(relx=0.4,rely=0.11, anchor=CENTER, relheight=0.03,relwidth=0.2)
        
        self.scrol.place(relx=0.96,rely=0.57, anchor=CENTER, relheight=0.7 ,relwidth=0.02) 
        self.salida.place(relx=0.57,rely=0.22, relheight=0.7,relwidth=0.38)
        self.borrar_salida.place(relx=0.95,rely=0.195, anchor="e")
        self.resultado.place(relx=0.57,rely=0.19, anchor="w", relheight=0.04,relwidth=0.07)
        self.sin_intervalo()


    def sin_intervalo(self):
        try:
            self.framecon.destroy()
        except: pass
        self.check_con_inter.bind("<Button-1>", lambda i:self.colorletracheck(2))
        self.check_sin_inter.bind("<Button-1>", lambda i: self.colorletracheck(1))
        self.framesin = Frame(self, background=self.colorlabel)
        self.labelsin = Label(self.framesin, font=("Segoe UI",round(widthroot*1.1/100)),
                             anchor="center",
                             background=self.colorlabel,
                             foreground=self.letra)
        self.scrol2 = Scrollbar(self.framesin, cursor="hand2",orient= VERTICAL,activebackground="gold2")
        self.opciones = ["Tabla de frecuencias","Tabla de dispersíon","Calcular intervalo","Moda", "Media", "Mediana", "Cuartiles","Varianza","Desviación estandar","Coeficiente de variación","Diagrama de caja","Polígonos de frecuencias","Ojiva","Rango"]
        self.valor_op = StringVar()
        self.combo_opciones = Combobox(self.framesin,font=("Segoe UI",round(widthroot*1.1/100)),
                                       values=self.opciones,
                                       state="readonly",
                                       textvariable=self.valor_op,
                                       cursor="hand2",
                                       background="black",
                                       foreground="gray9")
        self.combo_opciones.current(0)
        self.operacion = Label(self.framesin, text="Operación:",
                               fg=self.letra,
                               background=self.colorlabel,
                               font=("Segoe UI",round(widthroot*1.1/100)))
        self.labelsin.config(text="- Introduzca los datos separados por una coma ' , ' (solo númericos)\n- La representación decimal de los datos es mediante el punto")
        self.datos = Label(self.framesin, text="Datos:",
                           fg=self.letra,
                           font=("Segoe UI",round(widthroot*1.1/100)),
                           background=self.colorlabel)
        self.borrar_entrada = Button(self.framesin,text="Borrar",
                                     cursor="hand2",
                                     command= lambda :self.entrada_datos.delete(1.0,END))
        self.cargararch = Button(self.framesin,text="Cargar",
                                 cursor="hand2",
                                 command=self.cargarA)
        self.entrada_datos = Text(self.framesin,font=("Segoe UI",round(widthroot*1.1/100)),
                                  background=self.fondotext,
                                  foreground=self.letratext,
                                  yscrollcommand=self.scrol2.set)
        self.calcular = Button(self.framesin,text="Calcular",
                               font=("Segoe UI",round(widthroot*1.1/100),"bold"),
                               cursor="hand2",
                               foreground="gray9",
                               background="gold2",
                               borderwidth=5,
                               command=self.crear_operaciones)
        self.entrada_datos.focus()
        self.scrol2.config(command=self.entrada_datos.yview)            
        self.creador_Rboton(self.framesin,0,0)
        self.framesin.place(relx=0,rely=0.115,relheight=0.98,relwidth=0.55)
        self.entrada_datos.place(relx=0.5,rely=0.25, anchor=CENTER, relheight=0.21,relwidth=0.55)
        self.scrol2.place(relx=0.79,rely=0.25, anchor=CENTER, relheight=0.21,relwidth=0.03)
        self.borrar_entrada.place(relx=0.74,rely=0.12, anchor=CENTER)
        self.cargararch.place(relx=0.65,rely=0.12, anchor=CENTER)
        self.datos.place(relx=0.215,rely=0.12, anchor="w", relheight=0.04,relwidth=0.1)
        self.combo_opciones.place(relx=0.65,rely=0.53, anchor="e", relheight=0.05,relwidth=0.28)
        self.operacion.place(relx=0.22,rely=0.53, anchor="w", relheight=0.04,relwidth=0.13)
        self.calcular.place(relx=0.67,rely=0.507, relheight=0.05,relwidth=0.13)
        self.labelsin.place(relx=0,rely=0, relheight=0.07,relwidth=1)
               
        self.entrada_datos.delete(1.0,END)
        self.entrada_datos.insert(1.0,"1,2,3,4,5")
        self.boton_volver() 

    def agregar_fila(self):
        if len(self.intervalos_entries) <7:
            # Crear las celdas de intervalo y frecuencia para la nueva fila
            intervalo_entry = Entry(self.framecon, width=20)
            intervalo_entry.grid(row=len(self.intervalos_entries) + 7, padx = 5, column=0, pady=5)
            self.intervalos_entries.append(intervalo_entry)

            frecuencia_entry = Entry(self.framecon, width=10)
            frecuencia_entry.grid(row=len(self.frecuencias_entries) + 7, padx = 5, column=1, pady=5)
            self.frecuencias_entries.append(frecuencia_entry)
        else:
            if len(self.intervalos_entries) <14:
            # Crear las celdas de intervalo y frecuencia para la nueva fila
                intervalo_entry = Entry(self.framecon, width=20)
                intervalo_entry.grid(row=len(self.intervalos_entries)-7 + 9, column=2, padx=5, pady=5)
                self.intervalos_entries.append(intervalo_entry)

                frecuencia_entry = Entry(self.framecon, width=10)
                frecuencia_entry.grid(row=len(self.frecuencias_entries)-7 + 9, column=3, padx=5, pady=5)
                self.frecuencias_entries.append(frecuencia_entry)
                self.intervalo_label1.grid(row=8, column=2, pady=5)
                self.frecuenciaL1.grid(row=8, column=3, pady=5)
            else:
                if len(self.intervalos_entries) <21:
                    intervalo_entry = Entry(self.framecon, width=20)
                    intervalo_entry.grid(row=len(self.intervalos_entries)-14 + 9, column=4, padx=5, pady=5)
                    self.intervalos_entries.append(intervalo_entry)

                    frecuencia_entry = Entry(self.framecon, width=10)
                    frecuencia_entry.grid(row=len(self.frecuencias_entries)-14 + 9, column=5, padx=5, pady=5)
                    self.frecuencias_entries.append(frecuencia_entry)
                    self.intervalo_label2.grid(row=8, column=4, pady=5)
                    self.frecuenciaL2.grid(row=8, column=5, pady=5)

    def generar_intervalos(self, texto):
        
        def insertar_intevalos():
            self.borrar_entries()
            listaFi = []
            listaFi1= []
            contador = 4.0
            for j in range(k):
                for i in self.salida.get(contador,contador+1).strip().split(",()[]"):
                    s = i.strip(",[]()")
                    numeros = s.split(",")
                    for numero in numeros:
                        listaFi.append(float(numero))
                    listaFi1.append(listaFi)
                    listaFi = []
                    if j > 6 :
                        self.agregar_fila()
                        self.intervalos_entries[j].insert(j, str(i))
                    else:
                        self.intervalos_entries[j].insert(j, str(i))
                contador = contador+1.0
            Fi = []
            for i in self.entrada_sacarI.get(1.0,END).strip().split(","):
                Fi.append(float(i))

            contadorFi = 0
            conta = 0
            for j in listaFi1:
                for i in Fi:
                    if j == listaFi1[-1]:
                        if i >= j[0] and i <= j[1]:
                            contadorFi+=1
                    else:
                        if i >= j[0] and i < j[1]:
                            contadorFi+=1

                self.frecuencias_entries[conta].insert(0, contadorFi)
                contadorFi = 0
                conta+=1
                        

        try:
            numeros_texto = texto.get("1.0", "end").strip().split(",")
            numeros = [float(numero) for numero in numeros_texto]
        except:
            showerror(title="ERROR", message="Por favor verifique la integridad de los datos.\nDebe haber un caracter de mas.")
        minimo = min(numeros)
        maximo = max(numeros)
        rango = maximo - minimo
        k = round(1 + 3.322 * math.log10(len(numeros)))
        # Fórmula de Sturges
        amplitud = (rango / k)
        intervalos = []
        limite_inferior = minimo
        for i in range(k):
            limite_superior = round(limite_inferior + amplitud,2)
            if i+1 == k:
                intervalos.append((round(limite_inferior,2), maximo))
            else: 
                intervalos.append((round(limite_inferior,2), limite_superior))
            limite_inferior = limite_superior
        for i in reversed(intervalos):
            if intervalos[-1] == i :
                self.salida.insert(1.0, "["+str(i[0])+","+str(i[1])+"]\n")
            else:
                self.salida.insert(1.0,"["+str(i[0])+","+str(i[1])+")\n")
        self.salida.insert(1.0,F"AMPLITUD: {amplitud}\n\n")
        self.salida.insert(1.0,F"INTERVALOS: {k}\n")
        try:
            insertar_intevalos()
        except:pass

    def borrar_entries(self):
        for i in self.intervalos_entries:
            i.delete(0,END)
        for i in self.frecuencias_entries:
            i.delete(0,END)

    def borrar_datos_intervalos(self):
        self.entrada_sacarI.delete(1.0, END)
        for i in self.intervalos_entries:
            i.delete(0,END)
        for i in self.frecuencias_entries:
            i.delete(0,END)
        
    def con_intervalo(self):
        try:
            self.framesin.destroy()
        except:pass
        self.framecon = Frame(self, background=self.colorlabel)
        self.framecon.place(relx=0,rely=0.4,relheight=0.5,relwidth=0.55)
        separ = Frame(self, background="gray20").place(relx=0, rely=0.33, relheight=0.001, relwidth=0.55)
        self.intervalo_label1 = Label(self.framecon, text="Intervalos:",background=self.colorlabel,foreground=self.letra)
        self.frecuenciaL1 = Label(self.framecon, text="Frecuencia A",background=self.colorlabel,foreground=self.letra)
        self.intervalo_label2 = Label(self.framecon, text="Intervalos:",background=self.colorlabel,foreground=self.letra)
        self.frecuenciaL2 = Label(self.framecon, text="Frecuencia A",background=self.colorlabel,foreground=self.letra)
        #SACAR INTERVALO
        self.label_sacarI= Label(self, text= "GENERAR INTERVALO AUTOMATICAMENTE:\n-Ingrese los numeros separados por una coma ' , '",
                            font=("Segoe UI",round(widthroot*1/100)),
                             anchor="center",
                             background=self.colorlabel,
                             foreground=self.letra)
        self.label_sacarI.place(relx = 0.12, rely=0.113)
        scrol= Scrollbar(self, cursor="hand2",orient= VERTICAL,activebackground="gold2")
        self.entrada_sacarI= Text(self,font=("Segoe UI",round(widthroot*1.1/100)),
                                  background=self.fondotext,
                                  foreground=self.letratext,
                                  yscrollcommand=scrol.set)
        self.entrada_sacarI.focus()
        scrol.config(command=self.entrada_sacarI.yview)
        self.entrada_sacarI.place(relx = 0.08, rely=0.18, relheight=0.13, relwidth=0.35)
        scrol.place(relx=0.44,rely=0.2447, anchor=CENTER, relheight=0.13,relwidth=0.02)
        boton_borrar_sacarI = Button(self, text="Borrar", command=self.borrar_datos_intervalos).place(relx = 0.48, rely=0.22, anchor=CENTER)
        boton_sacarI = Button(self, text="Procesar", command=lambda:self.generar_intervalos(self.entrada_sacarI)).place(relx = 0.48, rely=0.28, anchor=CENTER)
        self.entrada_sacarI.insert(1.0, "1,2,3,4,5")
         # Crear las listas de Entry para los intervalos y frecuencias
        self.intervalos_entries = []
        self.frecuencias_entries = []
        # Crear las etiquetas de las columnas
        self.intervalo_label = Label(self.framecon, text="Intervalos:", background=self.colorlabel,foreground=self.letra)
        self.intervalo_label.grid(row=8, column=0, pady=5)
        self.frecuenciaL = Label(self.framecon, text="Frecuencia A",background=self.colorlabel,foreground=self.letra)
        self.frecuenciaL.grid(row=8, column=1, pady=5)

        # Crear el botón para agregar una nueva fila
        self.agregar_button = Button(self.framecon, text="Agregar fila", command=self.agregar_fila)
        self.agregar_button.grid(row=25, column=0)
        self.borrarI = Button(self.framecon, text="Borrar", command=self. borrar_datos_intervalos)
        self.borrarI.grid(row=25, column=1)


        # Crear las filas iniciales de la tabla
        for i in range(7):
            intervalo_entry = Entry(self.framecon, width=20)
            intervalo_entry.grid(row=i + 9, column=0, pady=5, padx=5)
            self.intervalos_entries.append(intervalo_entry)

            frecuencia_entry = Entry(self.framecon, width=10)
            frecuencia_entry.grid(row=i + 9, column=1, pady=5, padx=5)
            self.frecuencias_entries.append(frecuencia_entry)
        self.opciones = ["Tabla de frecuencias","Tabla de dispersíon","Moda", "Media", "Mediana", "Cuartiles","Varianza","Desviación estandar","Coeficiente de variación","Rango"]
        
        self.labelcon = Label(self, font=("Segoe UI",round(widthroot*1.1/100)),
                             anchor="center",
                             background=self.colorlabel,
                             foreground=self.letra)
        self.valor_op = StringVar()
        self.combo_opciones = Combobox(self.framecon,font=("Segoe UI",round(widthroot*1.1/100)),
                                       values=self.opciones,
                                       state="readonly",
                                       textvariable=self.valor_op,
                                       cursor="hand2",
                                       background="black",
                                       foreground="gray9")
        self.combo_opciones.current(0)
        self.operacion = Label(self.framecon, text="Operación:",
                               fg=self.letra,
                               background=self.colorlabel,
                               font=("Segoe UI",round(widthroot*1.1/100)))
        self.labelcon.config(text="- Ingrese los intervalos entre '[ )', '( ]', '( ), '[ ]' separados por una coma ' , '")
        self.calcular = Button(self.framecon,text="Calcular",
                               font=("Segoe UI",round(widthroot*1.1/100),"bold"),
                               cursor="hand2",
                               foreground="gray9",
                               background="gold2",
                               borderwidth=5,
                               command=self.crear_operaciones)
        #POBLACION O MUESTRA
        self.creador_Rboton(self.framecon,0.22, 0.785)
        self.combo_opciones.place(relx=0.69,rely=0.885, anchor="e", relheight=0.07,relwidth=0.42)
        self.operacion.place(relx=0.13,rely=0.88, anchor="w", relheight=0.06,relwidth=0.13)
        self.calcular.place(relx=0.71,rely=0.848, relheight=0.08,relwidth=0.12)
        self.labelcon.place(relx = 0.05, rely=0.34)
        

    def limpiar_entrada(self):
        funcion_sub_esp = re.sub("\n", "", self.capturar_datos)
        funcion_sub_let = re.sub("[a-z]", "", funcion_sub_esp)
        funcion_sub_car = re.sub("[*-+]", "", funcion_sub_let)
        funcion_sub_esp2 = re.sub(" ", "", funcion_sub_car)
        funcion_sub_comp = re.sub(",", " ", funcion_sub_esp2)
        entradalimpia =  funcion_sub_comp + " "
        verif = []
        variabletemp = ""
        for i in range(len(entradalimpia)):
            if entradalimpia[-2] == "." or entradalimpia[-2] =="*" or entradalimpia[-2] =="/":
                    showerror(title="ERROR", message="Verifique la integridad de los datos, puede haber un caracter de mas")
                    self.error_ = False
                    break
            try:
                if entradalimpia[i]=="0" or entradalimpia[i]=="1" or entradalimpia[i]=="2" or entradalimpia[i]=="3" or entradalimpia[i]=="4" or entradalimpia[i]=="5" or entradalimpia[i]=="6" or entradalimpia[i]=="7" or entradalimpia[i]=="8" or entradalimpia[i]=="9":
                    variabletemp = variabletemp + entradalimpia[i]
                elif entradalimpia[i] == ".":
                    variabletemp = variabletemp + entradalimpia[i]
                else:
                    verif.append(float(variabletemp))
                    variabletemp = ""
                self.error_ =  True
            except:
                showerror(title="ERROR", message="Verifique la integridad de los datos, puede haber un caracter de mas")
                self.error_ = False
                break

        for i in verif:
            try:
                self.arreglosin = np.append(self.arreglosin, i)
            except:pass
    def encontrar_intervalo_mas_repetido(self, datos):
        intervalos = []
        for intervalo in datos:
            inicio, fin = intervalo.split(":")
            intervalos.append((float(inicio), float(fin)))

        contador = Counter(intervalos)
        intervalo_mas_repetido = max(contador, key=contador.get)
        intervalo_mas_repetido = f"{intervalo_mas_repetido[0]}:{intervalo_mas_repetido[1]}"
        return intervalo_mas_repetido
    
    def limpiar_entradaRango(self):
        self.datos = self.entrada_datos.get("1.0", END).strip().split(",")
    
        if not self.datos:
            messagebox.showwarning("Advertencia", "Por favor, ingrese al menos un intervalo de datos.")
            return
    
        self.valores = []
        for intervalo in self.datos:
            try:
                inicio, fin = intervalo.split(":")
                inicio = float(inicio.strip())
                fin = float(fin.strip())
                self.valores.extend([valor for valor in range(int(inicio), int(fin) + 1)])
            except ValueError:
                messagebox.showwarning("Advertencia", f"El intervalo '{intervalo}' es inválido. Asegúrese de ingresar los valores correctamente.")
                return
        
        if not self.valores:
            messagebox.showwarning("Advertencia", "No se encontraron datos numéricos válidos.")
            return


    def crear_ventana_frecuencias(self, text_datos):
        def calcular_frecuencias():
            # Obtener los datos ingresados por el usuario
            datos = text_datos 
            if not datos:
                messagebox.showwarning("Advertencia", "Por favor, ingrese al menos un número.")
                return
            # Ordenar los datos de menor a mayor
            datos = sorted(datos, key=float)
            # Contar la frecuencia de cada número
            contador = Counter(datos)
            # Limpiar la tabla de frecuencias
            for widget in tabla_frecuencias.winfo_children():
                widget.destroy()

            # Crear encabezados de la tabla
            Label(tabla_frecuencias, text="Número(Xi)", relief=RIDGE).grid(row=0, column=0, sticky="nsew")
            Label(tabla_frecuencias, text="Frecuencia Absoluta(fi)", relief=RIDGE).grid(row=0, column=1, sticky="nsew")
            Label(tabla_frecuencias, text="Frecuencia Absoluta Acumulada(Fi)", relief=RIDGE).grid(row=0, column=2, sticky="nsew")
            Label(tabla_frecuencias, text="Frecuencia Relativa(hi)", relief=RIDGE).grid(row=0, column=3, sticky="nsew")
            Label(tabla_frecuencias, text="Frecuencia Relativa Acumulada(Hi)", relief=RIDGE).grid(row=0, column=4, sticky="nsew")
            Label(tabla_frecuencias, text="Porcentaje(%)", relief=RIDGE).grid(row=0, column=5, sticky="nsew")

            # Calcular totales
            total_datos = len(datos)
            total_frecuencia_acumulada = 0

            # Mostrar los resultados en la tabla de frecuencias
            row = 1
            for numero, frecuencia in contador.items():
                frecuencia_absoluta = frecuencia
                frecuencia_absoluta_acumulada = total_frecuencia_acumulada + frecuencia
                frecuencia_relativa = frecuencia / total_datos
                frecuencia_relativa_acumulada = frecuencia_absoluta_acumulada / total_datos
                porcentaje = frecuencia_relativa * 100

                Label(tabla_frecuencias, text=str(numero), relief=RIDGE).grid(row=row, column=0, sticky="nsew")
                Label(tabla_frecuencias, text=str(frecuencia_absoluta), relief=RIDGE).grid(row=row, column=1, sticky="nsew")
                Label(tabla_frecuencias, text=str(frecuencia_absoluta_acumulada), relief=RIDGE).grid(row=row, column=2, sticky="nsew")
                Label(tabla_frecuencias, text=str(round(frecuencia_relativa, 4)), relief=RIDGE).grid(row=row, column=3, sticky="nsew")
                Label(tabla_frecuencias, text=str(round(frecuencia_relativa_acumulada, 4)), relief=RIDGE).grid(row=row, column=4, sticky="nsew")
                Label(tabla_frecuencias, text=str(round(porcentaje, 2)) + "%", relief=RIDGE).grid(row=row, column=5, sticky="nsew")

                row += 1
                total_frecuencia_acumulada += frecuencia

        def copiar_tabla():
            # Obtener los datos de la tabla
            datos_tabla = []
            for row in range(1, tabla_frecuencias.grid_size()[1]):
                fila = []
                for col in range(6):
                    widget = tabla_frecuencias.grid_slaves(row=row, column=col)[0]
                    fila.append(widget.cget("text"))
                datos_tabla.append(fila)

            # Convertir los datos a una cadena de texto
            tabla_texto = ""
            for fila in datos_tabla:
                tabla_texto += "\t".join(fila) + "\n"

            # Copiar la tabla al portapapeles
            ventana_frecuencias.clipboard_clear()
            ventana_frecuencias.clipboard_append(tabla_texto)

        def crear_excel():
            # Obtener los datos de la tabla
            datos_tabla = []
            for row in range(1, tabla_frecuencias.grid_size()[1]):
                fila = []
                for col in range(6):
                    widget = tabla_frecuencias.grid_slaves(row=row, column=col)[0]
                    fila.append(widget.cget("text"))
                datos_tabla.append(fila)

            # Obtener los nombres de las columnas de la tabla
            nombres_columnas = []
            for col in range(6):
                widget = tabla_frecuencias.grid_slaves(row=0, column=col)[0]
                nombres_columnas.append(widget.cget("text"))

            # Crear el archivo Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Escribir los nombres de las columnas en la primera fila del archivo Excel
            for col, nombre in enumerate(nombres_columnas):
                sheet.cell(row=1, column=col + 1, value=nombre)

            # Escribir los datos en el archivo Excel
            for row, fila in enumerate(datos_tabla):
                for col, valor in enumerate(fila):
                    sheet.cell(row=row + 2, column=col + 1, value=valor)

            # Guardar el archivo Excel
            workbook.save("tabla_frecuencias.xlsx")
            messagebox.showinfo("Éxito", "Se ha creado el archivo Excel 'tabla_frecuencias.xlsx'.")

        # Crear la ventana de nivel superior
        ventana_frecuencias = Tk()
        ventana_frecuencias.title("Calculadora de Frecuencias")
        ventana_frecuencias.iconbitmap("img/icono.ico")
        ventana_frecuencias.resizable(0, 0)
        xr_ventana = ventana_frecuencias.winfo_screenwidth() // 2 - 600 // 2
        yr_ventana = ventana_frecuencias.winfo_screenheight() // 2 - 185 // 2
        ventana_frecuencias.geometry("" + "+" + str(xr_ventana) + "+" + str(yr_ventana))

        # Botón para copiar la tabla
        boton_copiar = Button(ventana_frecuencias, text="Copiar tabla", command=copiar_tabla, cursor="hand2")
        boton_copiar.grid(row=1,column=0,sticky=E)

        # Botón para crear el archivo Excel
        boton_excel = Button(ventana_frecuencias, text="Crear Excel", command=crear_excel, cursor="hand2")
        boton_excel.grid(row=1,column=1,sticky=W, padx=10)
        # Tabla de frecuencias
        tabla_frecuencias = Frame(ventana_frecuencias, relief=RIDGE, borderwidth=1)
        tabla_frecuencias.grid(row=2,columnspan=2)

        # Configurar el tamaño de las columnas de la tabla
        tabla_frecuencias.grid_columnconfigure(0, weight=1)
        tabla_frecuencias.grid_columnconfigure(1, weight=1)
        tabla_frecuencias.grid_columnconfigure(2, weight=1)
        tabla_frecuencias.grid_columnconfigure(3, weight=1)
        tabla_frecuencias.grid_columnconfigure(4, weight=1)
        tabla_frecuencias.grid_columnconfigure(5, weight=1)
        # Llamar calcular frecuencias
        calcular_frecuencias()
        # Ejecutar el bucle principal de la interfaz gráfica
        ventana_frecuencias.mainloop()


    def generar_poligono_frecuencias(self, text_datos):
        # Obtener los datos del widget Text
        datos = text_datos.get("1.0", "end-1c")

        # Verificar si se ingresaron datos
        if not datos:
            messagebox.showwarning("Advertencia", "Por favor, ingrese al menos un número.")
            return

        # Convertir los datos a una lista de números
        try:
            datos_numeros = [float(dato.strip()) for dato in datos.split(",")]
        except ValueError:
            messagebox.showwarning("Advertencia", "Los datos ingresados no son válidos.")
            return

        # Calcular la frecuencia de cada dato
        datos_unicos, frecuencias = np.unique(datos_numeros, return_counts=True)

        # Crear el polígono de frecuencias
        plt.plot(datos_unicos, frecuencias, marker='o')

        # Personalizar el polígono de frecuencias
        plt.title("Polígono de Frecuencias")
        plt.xlabel("Datos")
        plt.ylabel("Frecuencia")

        # Mostrar el polígono de frecuencias
        plt.show()

    def generar_ojiva(self, text_datos):
        # Obtener los datos del widget Text
        datos = text_datos.get("1.0", "end-1c")

        # Verificar si se ingresaron datos
        if not datos:
            messagebox.showwarning("Advertencia", "Por favor, ingrese al menos un número.")
            return

        # Convertir los datos a una lista de números
        try:
            datos_numeros = [float(dato.strip()) for dato in datos.split(",")]
        except ValueError:
            messagebox.showwarning("Advertencia", "Los datos ingresados no son válidos.")
            return

        # Ordenar los datos de menor a mayor
        datos_ordenados = np.sort(datos_numeros)

        # Calcular la frecuencia acumulada
        frecuencia_acumulada = np.arange(1, len(datos_ordenados) + 1) / len(datos_ordenados)

        # Crear la ojiva
        plt.plot(datos_ordenados, frecuencia_acumulada, marker='o')

        # Personalizar la ojiva
        plt.title("Ojiva")
        plt.xlabel("Datos")
        plt.ylabel("Frecuencia Acumulada")

        # Mostrar la ojiva
        plt.show()


    def generar_boxplot(self, text_datos):
        # Obtener los datos del widget Text
        datos = text_datos.get("1.0", "end-1c")

        # Verificar si se ingresaron datos
        if not datos:
            messagebox.showwarning("Advertencia", "Por favor, ingrese al menos un número.")
            return

        # Convertir los datos a una lista de números
        try:
            datos_numeros = [float(dato.strip()) for dato in datos.split(",")]
        except ValueError:
            messagebox.showwarning("Advertencia", "Los datos ingresados no son válidos.")
            return

        # Crear el gráfico de boxplot
        fig, ax = plt.subplots()

        # Configurar el boxplot con los datos
        boxplot = ax.boxplot(datos_numeros)

        # Personalizar el gráfico
        ax.set_title("Boxplot de Datos")
        ax.set_xlabel("Datos")
        ax.set_ylabel("Valores")

        # Agregar interactividad
        def on_click(event):
            # Cambiar el rango de los ejes
            if event.button == 'up':
                ax.set_ylim(ax.get_ylim()[0] - 1, ax.get_ylim()[1] + 1)
            elif event.button == 'down':
                ax.set_ylim(ax.get_ylim()[0] + 1, ax.get_ylim()[1] - 1)
            
            # Actualizar el gráfico
            plt.draw()

        def on_key(event):
            # Agregar etiquetas a los outliers
            if event.key == 'o':
                for line in boxplot['fliers']:
                    line.set_label('Outlier')
                plt.legend()
                plt.draw()

        # Conectar los eventos de interacción con el gráfico
        fig.canvas.mpl_connect('scroll_event', on_click)
        fig.canvas.mpl_connect('key_press_event', on_key)

        # Mostrar el gráfico
        plt.show()


    def preparar_intervalos(self, operacion):
        Li = np.array([])
        Ls = np.array([])
        self.xi = np.array([])
        fi = np.array([])
        A = np.array([])
        interM = np.array([])
        Fi = np.array([])
        datos = self.obtener_datos()
        # Calcular la frecuencia total
        frecuencia_total = sum(float(frecuencia) for intervalo, frecuencia in datos)
        N = frecuencia_total

        for intervalo, frecuencia in datos:
                inicio, fin = re.findall(r"\d+\.\d+|\d+", intervalo)
                inicio, fin = float(inicio), float(fin)
                interM = np.append(interM, intervalo)
                Li= np.append(Li,float(inicio))
                Ls = np.append(Ls, float(fin))
                self.xi= np.append(self.xi, round(float((inicio + fin) / 2),2))      
                fi = np.append(fi, round(float(frecuencia),2))
                A = np.append(A, fin - inicio)
        for i in range(len(fi)):
            if i == 0:
                Fi = np.append(Fi, round(float(fi[i]),2))
            else:
                Fi = np.append(Fi, round(float(fi[i]),2)+Fi[i-1])

        #MEDIA
        media= round(float(sum(self.xi * fi)/N),3)

        #MODA
        fi_menos1 = 0
        fi_mas1 = 0
        try:
            fi_menos1 = fi[fi.argmax()-1]
        except:pass
        try:
            fi_mas1 = fi[fi.argmax()+1]
        except:pass
        frecuencia_mas_alta = fi.max()

        posicion_frecuencia_mas_alta = np.where(fi == frecuencia_mas_alta)[0][0]
        moda = round(float(Li[posicion_frecuencia_mas_alta]+A[posicion_frecuencia_mas_alta]*((frecuencia_mas_alta - fi_menos1)/(frecuencia_mas_alta - fi_menos1 + frecuencia_mas_alta-fi_mas1))),2)
        intervalo_modal = interM[posicion_frecuencia_mas_alta]
        #MEDIANA
        intervalo_mediana = 0
        Fi_menos1 = 0
        for i in Fi:
            if i > N/2:
                intervalo_mediana = i
                break
        posicion = np.where(Fi == intervalo_mediana)[0][0]
        try:
            if posicion-1 >= 0:
                Fi_menos1 = Fi[posicion-1]
        except:pass
        mediana = round(float(Li[posicion] + A[posicion] * ((N/2-Fi_menos1)/fi[posicion])),3)

        #RANGO
        rango = round(float(Ls.max()-Li.min()),3)

        #CUARTILES
        k3 = (3*N)/4 
        k1 = (1*N)/4
        Fi_k3= 0
        Fi_k1 = 0
        Fk3_menos1 = 0
        Fk1_menos1 = 0
        for i in Fi:
            if i > k3:
                Fi_k3 = i
                break
        for i in Fi:
            if i > k1:
                Fi_k1 = i
                break
        posicion_k3= np.where(Fi == Fi_k3)[0][0]
        posicion_k1= np.where(Fi == Fi_k1)[0][0]
        try:
            if posicion_k3-1 >= 0:
                Fk3_menos1 = Fi[posicion_k3-1]
        except:pass
        try:
            if posicion_k1-1 >= 0:
                Fk1_menos1 = Fi[posicion_k1-1]
        except:pass
        
        q3 = float(Li[posicion_k3]+A[posicion_k3]*((k3-Fk3_menos1)/fi[posicion_k3]))
        q1 = float(Li[posicion_k1]+A[posicion_k1]*((k1-Fk1_menos1)/fi[posicion_k1]))
        
        #VARIANZA  
        varP = round(float(sum(((self.xi-media)**2)*fi)/N),3)
        varM = round(float(sum(((self.xi-media)**2)*fi)/(N-1)),3)
        
        #DESVIACION ESTANDAR
        desvP = np.sqrt(varP)
        desvM = np.sqrt(varM)
        
        #COEFICIENTE DE VARIACION
        cdvP = desvP/media
        cdvM = desvM/media

        if operacion == "media":      
            return media
        if operacion == "moda":
            return [moda, intervalo_modal]
        if operacion == "mediana":
            return mediana
        if operacion == "rango":
            return rango
        if operacion == "cuartiles":
            return [q1, q3]
        if operacion == "varianza":
            return [varP, varM]
        if operacion == "desviacion":
            return [desvP, desvM]
        if operacion == "coeficiente":
            return [cdvP, cdvM]
        if operacion == "mad":
            return self.xi
        if operacion == "tddP":
            return [media, moda, intervalo_modal,mediana,rango,q1,q3,varP,desvP,cdvP]
        if operacion == "tddM":
            return [media, moda, intervalo_modal,mediana,rango,q1,q3,varM,desvM,cdvM]
        
    def obtener_datos(self):
            datos = []
            for i in range(len(self.intervalos_entries)):
                intervalo = self.intervalos_entries[i].get()
                frecuencia = self.frecuencias_entries[i].get()

                # Verificar si la fila contiene datos en la primera columna
                if not intervalo:
                    break  # Salir del bucle si no hay datos en la primera columna

                # Validar el formato del intervalo
                if not re.match(r"^\[\d+(\.\d+)?,\d+(\.\d+)?\]$|^\(\d+(\.\d+)?,\d+(\.\d+)?\]$|^\[\d+(\.\d+)?,\d+(\.\d+)?\)$|^\(\d+(\.\d+)?,\d+(\.\d+)?\)$", intervalo):
                    messagebox.showerror("Error", "El intervalo debe tener el formato '(inicio, fin]', '[inicio, fin)', '[inicio, fin]', o '(inicio, fin)'.")
                    return []

                # Validar si la frecuencia absoluta es un número
                if not frecuencia.replace(".", "", 1).isdigit():
                    messagebox.showerror("Error", "La frecuencia absoluta debe ser un valor numérico.")
                    return []

                datos.append((intervalo, frecuencia))
            return datos

    def crear_tabla_datos(self):
        
    
        def copiar_datos(tree):
            # Obtener todos los items de la tabla
            items = tree.get_children()

            # Verificar si hay datos en la tabla
            if len(items) == 0:
                messagebox.showwarning("Advertencia", "No hay datos en la tabla para copiar.")
                return

            # Crear una lista para almacenar los datos de todas las filas
            datos_tabla = []

            # Iterar sobre los items y obtener los valores de cada columna
            for item in items:
                valores = []
                for column in tree["columns"]:
                    valor = tree.item(item, "values")[tree["columns"].index(column)]
                    valores.append(valor)
                datos_tabla.append(valores)

            # Crear una cadena de texto con los datos de la tabla
            texto = ""
            for fila in datos_tabla:
                texto += "\t".join(fila) + "\n"

            # Copiar los datos al portapapeles
            pyperclip.copy(texto)
            messagebox.showinfo("Copiado", "Los datos se han copiado al portapapeles.")

        def crear_excel(tree):
            # Obtener todos los items de la tabla
            items = tree.get_children()

            # Verificar si hay datos en la tabla
            if len(items) == 0:
                messagebox.showwarning("Advertencia", "No hay datos en la tabla para crear el archivo Excel.")
                return

            # Crear un DataFrame para almacenar los datos de la tabla
            datos_tabla = []

            # Obtener los nombres de las columnas
            columnas = tree["columns"]

            # Iterar sobre los items y obtener los valores de cada columna
            for item in items:
                valores = []
                for column in columnas:
                    valor = tree.item(item, "values")[columnas.index(column)]
                    valores.append(valor)
                datos_tabla.append(valores)

            # Crear un DataFrame con los datos y los nombres de las columnas
            df = pd.DataFrame(datos_tabla, columns=columnas)

            # Crear un archivo Excel y guardar los datos en la hoja de trabajo
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Escribir los nombres de las columnas en la primera fila
            for col_num, columna in enumerate(df.columns, 1):
                sheet.cell(row=1, column=col_num, value=columna)

            # Escribir los datos en las filas siguientes
            for row_num, fila in enumerate(df.itertuples(), 2):
                for col_num, valor in enumerate(fila[1:], 1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Abrir el cuadro de diálogo para guardar el archivo
            file_path = tk.filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

            # Guardar el archivo Excel
            if file_path:
                workbook.save(file_path)
                messagebox.showinfo("Excel creado", "Se ha creado el archivo Excel exitosamente.")

            

        def mostrar_datos():
            self.datos = self.obtener_datos()

            
            # Verificar si los campos de entrada están vacíos
            if not self.datos:
                return

            # Crear una ventana nueva para mostrar los datos
            ventana_resultado = tk.Toplevel(root)
            ventana_resultado.title("Calculadora de Frecuencias")
            ventana_resultado.iconbitmap("icono.ico")
            xr_ventana = ventana_resultado.winfo_screenwidth() // 2 - 600 // 2
            yr_ventana = ventana_resultado.winfo_screenheight() // 2 - 185 // 2
            ventana_resultado.geometry("" + "+" + str(xr_ventana) + "+" + str(yr_ventana))
            # Crear el Treeview para mostrar los datos en una tabla
            tree = ttk.Treeview(ventana_resultado, show="headings")  # Mostrar solo los encabezados

            # Definir las columnas y encabezados
            tree["columns"] = ("Intervalos", "Xi", "Frecuencia Absoluta", "Frecuencia Absoluta Acumulada", "Frecuencia Relativa", "Frecuencia Relativa Acumulada", "Porcentaje")
            tree.heading("Intervalos", text="Intervalos")
            tree.heading("Xi", text="Xi")
            tree.heading("Frecuencia Absoluta", text="Frecuencia Absoluta")
            tree.heading("Frecuencia Absoluta Acumulada", text="Frecuencia Absoluta Acumulada")  # Cambio de "Frecuencia Absoluta" a "Fi"
            tree.heading("Frecuencia Relativa", text="Frecuencia Relativa")
            tree.heading("Frecuencia Relativa Acumulada", text="Frecuencia Relativa Acumulada")
            tree.heading("Porcentaje", text="Porcentaje")


             # Crear un marco para los botones Copiar y Crear Excel
            # Crear un marco para los botones Copiar y Crear Excel
            botones_frame = tk.Frame(ventana_resultado)
            botones_frame.pack(side=tk.TOP, padx=10, pady=5)

            # Crear el botón Copiar en el marco de botones
            copiar_button = tk.Button(botones_frame, text="Copiar", command=lambda: copiar_datos(tree))
            copiar_button.pack(side=tk.LEFT, padx=5)

            # Crear el botón Crear Excel en el marco de botones
            excel_button = tk.Button(botones_frame, text="Crear Excel", command=lambda: crear_excel(tree))
            excel_button.pack(side=tk.LEFT, padx=5)


            # Mantener un conjunto para verificar duplicados
            datos_insertados = set()

            # Calcular la frecuencia total
            frecuencia_total = sum(float(frecuencia) for intervalo, frecuencia in self.datos)
            # Calcular la frecuencia absoluta acumulada
            frecuencia_acumulada = 0

            # Insertar los datos en la tabla
            for intervalo, frecuencia in self.datos:
                inifin = intervalo
                inicio, fin = re.findall(r"\d+\.\d+|\d+", intervalo)
                inicio, fin = float(inicio), float(fin)

                datos_unificados = f"{inifin[0]}{inicio}, {fin}{inifin[-1]}"
                xi = round(float((inicio + fin) / 2),2)
                frecuencia_absoluta = round(float(frecuencia),2)
                frecuencia_acumulada += round(float(frecuencia),2)
                frecuencia_relativa = round(float(frecuencia) / frecuencia_total,2)
                frecuencia_relativa_acumulada = round(float(frecuencia_acumulada / frecuencia_total),2)
                porcentaje = round(frecuencia_relativa * 100, 2)

              
            

                if datos_unificados not in datos_insertados:
                    tree.insert("", "end", values=(datos_unificados, xi, frecuencia_absoluta, frecuencia_acumulada, frecuencia_relativa, frecuencia_relativa_acumulada, f"{porcentaje}%"))
                    datos_insertados.add(datos_unificados)

            

            # Ajustar el ancho de las columnas
            tree.column("Intervalos", width=90)
            tree.column("Xi", width=90)
            tree.column("Frecuencia Absoluta", width=150)
            tree.column("Frecuencia Absoluta Acumulada", width=200)
            tree.column("Frecuencia Relativa", width=150)
            tree.column("Frecuencia Relativa Acumulada", width=200)
            tree.column("Porcentaje", width=100)

            tree.pack()


        mostrar_datos()

    def crear_operaciones(self):
        self.rangocomp= []
        self.arreglosin = np.array([])
        try:
            self.capturar_datos = self.entrada_datos.get(1.0,END)
        except:pass
        control = self.controlcheckB.get()
        combo = self.combo_opciones.get()
        self.error_ = False
        if re.search("[1-9]", self.salida.get(1.0, END)):
            self.salida.insert(1.0,"\n")
        else:pass
        if control == "sinI":
            if combo == "Varianza" or combo == "Desviación estandar" or combo =="Coeficiente de variación"or combo =="Tabla de dispersíon":
                    self.limpiar_entrada()
                    if self.error_ == True:
                        radioB = self.valor_r.get()
                        if combo == "Varianza":
                            #POBLACIÓN
                            if radioB == 1:
                                self.salida.insert(1.0, str("Varianza (σ2): "+ str(np.var(self.arreglosin))))
                            #MUESTRA 
                            else: self.salida.insert(1.0, str("Varianza (s2): "+ str(st.tvar(self.arreglosin))))
                        if combo == "Desviación estandar":
                            #POBLACIÓN
                            if radioB == 1:
                                self.salida.insert(1.0, str("Desviación estándar (σ2): "+ str(np.std(self.arreglosin))))
                            #MUESTRA 
                            else: self.salida.insert(1.0, str("Desviación estándar (s2): "+ str(stat.stdev(self.arreglosin))))
                        if combo == "Coeficiente de variación":
                            #POBLACIÓN
                            if radioB == 1:
                                self.salida.insert(1.0, str("Coeficiente de variación (σ2): "+ str(np.std(self.arreglosin)/self.arreglosin.mean())))
                            #MUESTRA 
                            else: self.salida.insert(1.0, str("Coeficiente de variación (s2): "+ str(stat.stdev(self.arreglosin)/self.arreglosin.mean())))
                        if combo == "Tabla de dispersíon":
                            q =stat.quantiles(self.arreglosin)
                            moda = st.mode(self.arreglosin).mode
                            modacantidad = st.mode(self.arreglosin).count
                            
                            #POBLACIÓN
                            if radioB == 1:
                                try:
                                    self.salida.insert(1.0 , "Tamaño de la población: "+str(self.arreglosin.size)+"\n"+
                                    "Q1: "+str(q[0])+"\n"+ "Q3: "+str(q[2])+"\n"+ "Rango intercuartilico: "+str(q[2]-q[0])+"\n"+"Desviación cuartil: "+ str((q[2]-q[0])/2)+"\n"+
                                    "Valor mas bajo: "+str(self.arreglosin.min())+"\n"+"Valor mas alto: "+str(self.arreglosin.max())+"\n"+ "Moda: "+ str(float(moda))+" - Cantidad de repeticiones: "+ str(float(modacantidad))+"\n"+
                                    "Media (μ): "+ str(round(self.arreglosin.mean(),3))+"\n"+
                                    "Mediana: "+str(round(np.median(self.arreglosin),3))+"\n"+
                                    "Rango: "+str((self.arreglosin.max()-self.arreglosin.min()))+"\n"+
                                    str("Varianza (σ2): "+ str(np.var(self.arreglosin)))+"\n"+
                                    str("Desviación estándar (σ2): "+ str(np.std(self.arreglosin)))+"\n"+               str("Coeficiente de variación (σ2): "+ str(np.std(self.arreglosin)/self.arreglosin.mean())))
                                except:pass
                            #MUESTRA 
                            else:
                                try:
                                    self.salida.insert(1.0 ,"Tamaño de la muestra: "+str(self.arreglosin.size)+"\n"+
                                    "Q1: "+str(q[0])+"\n"+ "Q3: "+str(q[2])+"\n"+ "Rango intercuartilico: "+str(q[2]-q[0])+"\n"+"Desviación cuartil: "+ str((q[2]-q[0])/2)+"\n"+
                                    "Valor mas bajo: "+str(self.arreglosin.min())+"\n"+
                                    "Valor mas alto: "+str(self.arreglosin.max())+"\n"+
                                    "Moda: "+ str(float(moda))+" - Cantidad de repeticiones: "+ str(float(modacantidad))+"\n"+
                                    "Media (x̄): "+ str(round(self.arreglosin.mean(),3))+"\n"+
                                    "Mediana: "+str(round(np.median(self.arreglosin),3))+"\n"+
                                    "Rango: "+str((self.arreglosin.max()-self.arreglosin.min()))+"\n"+
                                    str("Varianza (s2): "+ str(st.tvar(self.arreglosin)))+"\n"+
                                    str("Desviación estándar (s2): "+ str(stat.stdev(self.arreglosin)))+"\n"+
                                    str("Coeficiente de variación (s2): "+ str(stat.stdev(self.arreglosin)/self.arreglosin.mean())))
                                except:pass
                    else:pass


            else:
                    self.limpiar_entrada()
                    if self.error_ == True:
                        if combo == "Calcular intervalo":
                            self.generar_intervalos(self.entrada_datos)
                        if combo == "Tabla de frecuencias":
                            text_datos = self.entrada_datos.get(1.0, END).strip().split(",")
                            self.crear_ventana_frecuencias(text_datos)
                        if combo == "Cuartiles":
                            q =stat.quantiles(self.arreglosin)
                            self.salida.insert(1.0 ,"Q1: "+str(q[0])+"\n"+ "Q3: "+str(q[2])+"\n"+ "Rango intercuartilico: "+str(q[2]-q[0])+"\n"+"Desviación cuartil: "+ str((q[2]-q[0])/2))
                        if combo == "Moda":
                            try:
                                moda = st.mode(self.arreglosin).mode
                                modacantidad = st.mode(self.arreglosin).count
                                self.salida.insert(1.0, "Moda: "+ str(float(moda))+" - Cantidad de repeticiones: "+ str(float(modacantidad)))
                            except:pass
                        if combo =="Media":
                            self.salida.insert(1.0,"Media: "+ str(round(self.arreglosin.mean(),3)))
                        if combo =="Mediana":
                            self.salida.insert(1.0,"Mediana: "+str(round(np.median(self.arreglosin),3)))  
                        if combo =="Rango":
                            self.salida.insert(1.0,"Rango: "+str((self.arreglosin.max()-self.arreglosin.min())))
                        if combo =="Diagrama de caja":
                            self.generar_boxplot(self.entrada_datos)
                        if combo == "Polígonos de frecuencias":
                            self.generar_poligono_frecuencias(self.entrada_datos)
                        if combo == "Ojiva":
                            self.generar_ojiva(self.entrada_datos)
                    else:pass
        #CON INTERVALO!!!!!            
        if control == "conI":
            if combo == "Varianza" or combo == "Desviación estandar" or combo =="Coeficiente de variación"or combo =="Tabla de dispersíon":
                        radioB = self.valor_r.get()
                        if combo == "Varianza":
                            var = self.preparar_intervalos("varianza")
                            #POBLACIÓN
                            if radioB == 1:
                                self.salida.insert(1.0, str("Varianza (σ2): "+ str(var[0])+"\n"))
                            #MUESTRA 
                            else: self.salida.insert(1.0, str("Varianza (s2): "+ str(var[1])+"\n"))
                        if combo == "Desviación estandar":
                            desv = self.preparar_intervalos("desviacion")
                            #POBLACIÓN
                            if radioB == 1:
                                self.salida.insert(1.0, str("Desviación estándar (σ2): "+ str(desv[0])+"\n"))
                            #MUESTRA 
                            else: self.salida.insert(1.0, str("Desviación estándar (s2): "+ str(desv[1])+"\n"))
                        
                        if combo == "Coeficiente de variación":
                            cdv = self.preparar_intervalos("coeficiente")
                            #POBLACIÓN
                            if radioB == 1:
                                self.salida.insert(1.0, str("Coeficiente de variación (σ2): "+ str(cdv[0])+"\n"))
                            #MUESTRA 
                            else: self.salida.insert(1.0, str("Coeficiente de variación (s2): "+ str(cdv[1])+"\n"))
                        if combo == "Tabla de dispersíon":
                            tddP = self.preparar_intervalos("tddP")
                            tddM = self.preparar_intervalos("tddM")
                            #POBLACIÓN
                            if radioB == 1:
                                try:
                                    titulos = ["coeficiente de variación (σ2): ", "Desviacion estandar (σ2): ", "Varianza (σ2): ","Q3: ", "Q1: ", "Rango: ", "Mediana: ", "Intervalo modal: ", "Moda: ", "Media (x̄): "]
                                    contador = 0
                                    for i in reversed(tddP):
                                        self.salida.insert(1.0, titulos[contador]+ str(i)+"\n")
                                        contador+=1
                                except:pass
                            #MUESTRA 
                            else:
                                try:
                                    try:
                                        titulos = ["coeficiente de variación (s2): ", "Desviacion estandar (s2): ", "Varianza (s2): ","Q3: ", "Q1: ", "Rango: ", "Mediana: ", "Intervalo modal: ", "Moda: ", "Media (μ): "]
                                        contador = 0
                                        for i in reversed(tddM):
                                            self.salida.insert(1.0, titulos[contador]+ str(i)+"\n")
                                            contador+=1
                                    except:pass
                                    
                                except:pass

            else:
                if combo == "Tabla de frecuencias":
                    self.crear_tabla_datos()
                if combo == "Cuartiles":
                    cuartiles = self.preparar_intervalos("cuartiles")
                    self.salida.insert(1.0,"Q1: "+str(cuartiles[0])+"\n"+"Q3: "+str(cuartiles[1])+"\n")
                if combo == "Moda":
                    moda = self.preparar_intervalos("moda")
                    self.salida.insert(1.0,"Moda: "+ str(moda[0])+ "\n" +"Intervalo Modal: "+str(moda[1])+"\n")
                if combo =="Media":
                    media = self.preparar_intervalos("media")
                    self.salida.insert(1.0,"Media: "+ str(media)+"\n")
                if combo =="Mediana":
                    mediana = self.preparar_intervalos("mediana")
                    self.salida.insert(1.0,"Mediana: "+str(mediana)+"\n") 
                if combo =="Rango":
                    rango = self.preparar_intervalos("rango")
                    self.salida.insert(1.0,"Rango: "+str(rango)+"\n")

    
#funcion para compartir redes sociales en una ventana secundaria
def redes():
    ventana_redes = Toplevel(root)
    ventana_redes.iconbitmap("img/icono.ico")
    ventana_redes.resizable(0,0)
    xr_ventana = root.winfo_screenwidth() // 2 - 600// 2
    yr_ventana = root.winfo_screenheight() // 2 - 185 // 2
    ventana_redes.geometry("600x185" + "+" +str(xr_ventana) + "+" + str(yr_ventana))
    ventana_redes.title("Redes sociales: ")
    gmail = "mailto:karim.dev.tech@gmail.com"
    github = "https://github.com/KarimDevTech"
    linkedin = "https://www.linkedin.com/in/marcelokarimjurigaray/"
    fondoGlabel = Label(ventana_redes, image= fondo_gmail,).grid(row=0, column=0)
    fondoGilabel =Label(ventana_redes, image= fondo_github).grid(row=1, column=0, pady=10)
    fondoLilabel = Label(ventana_redes, image= fondo_linkedin).grid(row=2, column=0)
    label_gmail = Label(ventana_redes,text= "karim.dev.tech@gmail.com", font=("Segoe UI", 16 , "underline bold"), fg="blue", cursor="hand2")
    label_gmail.grid(row=0, column=1,sticky=W)
    label_gmail.bind("<Button-1>", lambda i:webbrowser.open(gmail))

    label_github = Label(ventana_redes,text= github, font=("Segoe UI", 16 , "underline bold"), fg="blue", cursor="hand2")
    label_github.grid(row=1, column=1, pady=10,sticky=W)
    label_github.bind("<Button-1>", lambda i:webbrowser.open(github))

    label_linkedin = Label(ventana_redes,text= linkedin, font=("Segoe UI", 16 , "underline bold"), fg="blue", cursor="hand2")
    label_linkedin.grid(row=2, column=1, sticky=W)
    label_linkedin.bind("<Button-1>", lambda i:webbrowser.open(linkedin))
    

#CONFIGURACIÓN ROOT  
root = Tk()
root.iconbitmap("img/icono.ico")
root.resizable(0,0)
monitorH = root.winfo_screenheight()
monitorW =root.winfo_screenwidth()
widthroot= round(monitorW/1.5)
height= round(monitorH/1.5)
x_ventana = root.winfo_screenwidth() // 2 - widthroot// 2
y_ventana = root.winfo_screenheight() // 2 - height // 2
posicion = str(widthroot) + "x" + str(height) + "+" + str(x_ventana) + "+" + str(y_ventana)
root.geometry(posicion)
root.title("STADISKIS - Calculadora automática de estadística")
aplicacion = app(root,widthroot, height)
aplicacion.place(relx=0,rely=0.03, relheight = 1, relwidth= 1)

#IMAGENES
fondo_gmail = ImageTk.PhotoImage(Image.open("img/gmail.png").resize((50,50)))
fondo_github = ImageTk.PhotoImage(Image.open("img/github.png").resize((50,50)))
fondo_linkedin =ImageTk.PhotoImage(Image.open("img/linkedin.png").resize((50,50)))


#MENU PARA LA APP
#Se crea un frame para contener el menu
menu_frame = Frame(root, background='gold4')
#Posiciona el Frame en la parte superior de la ventana
menu_frame.place(relx=0, rely=0, relheight=0.05, relwidth=1)
#Se crea los Menubutton dentro del Frame
archivo = Menubutton(menu_frame, 
                        text='Archivo', 
                        background='gold2', 
                        foreground='gray9', 
                        activeforeground='gray9', 
                        activebackground='gold3',
                        font=("Segoe UI",round(widthroot*1.1/100)),
                        cursor="hand2"
                        )
tema = Menubutton(menu_frame,
                    text='Tema', 
                    background='gold2', 
                    foreground='gray9', 
                    activeforeground='gray9', 
                    activebackground='gold3',
                    font=("Segoe UI",round(widthroot*1.1/100)),
                    cursor="hand2")
doc = Menubutton(menu_frame,
                    text='Documentación', 
                    background='gold2', 
                    foreground='gray9', 
                    activeforeground='gray9', 
                    activebackground='gold3',
                    font=("Segoe UI",round(widthroot*1.1/100)),
                    cursor="hand2")

#Se crean variables para almacenar opciones desplegable para el Menubutton
menu_archivo = Menu(archivo, tearoff=0)
menu_tema = Menu(tema, tearoff=0)
menu_doc = Menu(doc, tearoff=0)
carga_archivos = Menu(menu_archivo,tearoff=0)
#Se asigna el menú desplegable al Menubutton
archivo.config(menu=menu_archivo)
tema.config(menu=menu_tema)
doc.config(menu=menu_doc)
#funcion para cambio de tema
def cambio_tema(color):

    if color== 0:
        aplicacion.tema(0)
        menu_frame.config(background="gold4")
        archivo.config(background="gold2",
                       foreground="gray9",
                       activebackground="gold3",
                       activeforeground="gray9")
        tema.config(background="gold2",
                       foreground="gray9",
                       activebackground="gold3",
                       activeforeground="gray9")
        doc.config(background="gold2",
                       foreground="gray9",
                       activebackground="gold3",
                       activeforeground="gray9")
        menu_archivo.config(background="gray90",
                        foreground="gray9",
                        activebackground="gold2",
                       activeforeground="gray9")
        carga_archivos.config(background="gray90",
                        foreground="gray9",
                        activebackground="gold2",
                       activeforeground="gray9")
        menu_doc.config(background="gray90",
                        foreground="gray9",
                        activebackground="gold2",
                       activeforeground="gray9")
        menu_tema.config(background="gray90",
                        foreground="gray9",
                        activebackground="gold2",
                       activeforeground="gray9")
    else:
        aplicacion.tema(1)
        menu_frame.config(background="gray2")
        archivo.config(background="gray9",
                       foreground="gray50",
                       activebackground="gray25",
                       activeforeground="gray60")
        tema.config(background="gray9",
                       foreground="gray50",
                       activebackground="gray25",
                       activeforeground="gray60")
        doc.config(background="gray9",
                       foreground="gray50",
                       activebackground="gray25",
                       activeforeground="gray60")
        menu_archivo.config(background="gray12",
                       foreground="gray50")
        carga_archivos.config(background="gray12",
                       foreground="gray50")
        menu_doc.config(background="gray12",
                       foreground="gray50")
        menu_tema.config(background="gray12",
                       foreground="gray50")
cambio_tema(1)
#Se Agrega opciónes al menú desplegable 'Archivo'
menu_archivo.add_cascade(label='Cargar archivo', 
                         activeforeground='black', 
                         activebackground='gold2',
                         font=("Segoe UI",round(widthroot*1/100)),
                         menu=carga_archivos,
                         )
carga_archivos.add_command(label="Datos sin intervalo...",
                           command=aplicacion.cargarA,
                           activeforeground='black', 
                           activebackground='gold2',
                           font=("Segoe UI",round(widthroot*1/100)))
carga_archivos.add_command(label="Datos con intervalo...",
                            command=aplicacion.cargarB,
                           activeforeground='black',
                           activebackground='gold2',
                           font=("Segoe UI",round(widthroot*1/100)))
menu_archivo.add_separator()
menu_archivo.add_command(label='Salir', 
                         command=aplicacion.salir, 
                         activeforeground='gray9', 
                         activebackground='gray60',
                         font=("Segoe UI",round(widthroot*1/100)))

menu_doc.add_command(label="Acerca de Stadiskis",
                     command= lambda: showinfo("Acerca de Stadiskis:",
                                               "Esta interfaz fue diseñada por Marcelo Karim Juri Garay\n\nCopyright © 2023 Stadiskis Software. Reservados todos los derechos"), 
                         activeforeground='gray9', 
                         activebackground='gray60',
                         font=("Segoe UI",round(widthroot*1/100)))
menu_doc.add_separator()
menu_doc.add_command(label="Redes Sociales",
                    command=redes, 
                    activeforeground='gray9', 
                    activebackground='gray60',
                    font=("Segoe UI",round(widthroot*1/100)))
menu_tema.add_command(label='Oscuro', 
                         activeforeground='black', 
                         activebackground='gold2',
                         font=("Segoe UI",round(widthroot*1/100)),
                         command=lambda: cambio_tema(1))
menu_tema.add_command(label='Claro', 
                         activeforeground='black', 
                         activebackground='gold2',
                         font=("Segoe UI",round(widthroot*1/100)),
                         command=lambda: cambio_tema(0))
#Se posiciona los Menubutton dentro del Frame
archivo.place(relx=0.00 , rely=0, relheight=1, relwidth=0.07)
tema.place(relx=0.07, rely=0, relheight=1,relwidth=0.07)
doc.place(relx=0.14, rely=0, relheight=1, relwidth=0.12)



aplicacion.mainloop()

